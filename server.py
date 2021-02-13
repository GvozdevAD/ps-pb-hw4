from pandas import read_excel
from openpyxl import load_workbook
from collections import Counter, defaultdict

def count_browser_hits_month(browser, object_list, month):
    """Функция считает сколько раз заходили с определенного браузера в определенный месяц
    :param browser: имя браузера
    :param object_list: список 
    :param mounth: месяц в виде числа
    """
    result = 0
    for element in object_list:
        if element['Браузер'] == browser:
            if element['Дата посещения'].month == month:
                result += 1
    return result

def count_purch_items_by_month(product, object_list, month):
    """Функция считает сколько раз купили определенный товар в определенный месяц
    :param product: имя товара
    :param object_list: список 
    :param mounth: месяц в виде числа
    """
    result = 0
    for elements in logs_dict:
        goods = str(elements['Купленные товары']).split(',')
        for i in goods:
            if i == product:
                if elements['Дата посещения'].month == month:
                    result += 1
    return result

def gender_items(object_list, gender):
    """Функция считает сколько товаров купили мужчины или женщины
    :param object_list: Список
    :param gender: пол"""
    gender_dict = defaultdict(int)
    
    for elements in logs_dict:
        goods = str(elements['Купленные товары']).split(',')
        for products in goods:
            if products == 'Ещё 2 варианта' or products== 'Ещё 3 варианта':
                del products 
            elif elements['Пол'] == gender:
                gender_dict[products] += 1
    return gender_dict

excel_logs = read_excel('logs.xlsx', sheet_name='log', engine='openpyxl')
logs_dict = excel_logs.to_dict(orient='records')

#Создадим dict под топ браузеры 
top_browsers = defaultdict(int)
#Создадим dict под топ купленных товаров
goods_dict = defaultdict(int)

# Перебираем циклом все браузеры и добавляем в dict
for element in logs_dict:
    top_browsers[element['Браузер']] += 1

# Перебираем циклом все купленные  товары и добавляем в dict
for elements in logs_dict:
    goods = str(elements['Купленные товары']).split(',')
    for products in goods:
        goods_dict[products] += 1

# Удаляем не товары
del goods_dict['Ещё 2 варианта']
del goods_dict['Ещё 3 варианта']


top_browsers = Counter(top_browsers)
goods_dict_count = Counter(goods_dict)
goods_dict_count_m = Counter(gender_items(logs_dict, 'м'))
goods_dict_count_w = Counter(gender_items(logs_dict, 'ж'))

#Открываем для записи excel 
wb = load_workbook(filename='report.xlsx')
sheet = wb['Лист1']

# Записываем топ 7 браузеров и купленных товаров в excel
for row_1, row_2, i in zip(range(5,12),range(19,26), range(0,7)):
    sheet.cell(row=row_1, column=1).value = top_browsers.most_common(7)[i][0]
    sheet.cell(row=row_2, column=1).value = goods_dict_count.most_common(7)[i][0]

# Записываем сколько раз в месяц заходили с определенного браузера и купили определенных товаров
for row_1, row_2, i in zip(range(5,12), range(19,26), range(0,7)):
        for colum,m in zip(range(3,15), range(1,13)):
            sheet.cell(row=row_1, column=colum).value = count_browser_hits_month(top_browsers.most_common(7)[i][0], logs_dict, m)
            sheet.cell(row=row_2, column=colum).value = count_purch_items_by_month(goods_dict_count.most_common(7)[i][0], logs_dict, m)

# Записываем Предпочтения
sheet["B31"] = goods_dict_count_m.most_common(1)[0][0]
sheet["B32"] = goods_dict_count_m.most_common()[:-(len(goods_dict_count_m)+1):-1][0][0]
sheet["B33"] = goods_dict_count_w.most_common(1)[0][0]
sheet["B34"] = goods_dict_count_w.most_common()[:-(len(goods_dict_count_w)+1):-1][0][0]

# Сохраняем excel
wb.save('report.xlsx')
